from collections import deque
import openpyxl
import copy
from openpyxl.styles import PatternFill
import os

### 変数の定義 ###
# 変数や配列など
num_glass, num_color = map(int, input("num_glass, num_color: ").split())
glass = []
top_layer = []
space = []
ans = []
appeared = set()
# ワークブック、ワークシートの読み込み
wb = openpyxl.load_workbook("c:/Users/ochi yuma/OneDrive/ドキュメント/programming/Python/LoveWaterGlass.xlsm", keep_vba=True)
ws = wb["Glass"]

### 関数の定義 ###
def topLayerCheck(glass, index): # 一番上にある層の色と高さを取得
    if(len(glass[index]) == 0):
        return [-1, -1]
    t = glass[index].popleft()
    n = 1
    if(len(glass[index]) == 0):
        glass[index].appendleft(t)
        return [t, n]
    s = glass[index].popleft()
    while(s == t):
        n += 1
        if(len(glass[index])):
            s = glass[index].popleft()
        else:
            break
    if(t != s): glass[index].appendleft(s)
    for j in range(n): glass[index].appendleft(t)
    return [t, n]

def transferableJudge(top_layer, space, f, t): # fからtに色水を移せるか確認
    if(((top_layer[f][0] == top_layer[t][0]) or (top_layer[t][0] == -1)) and (space[t] - top_layer[f][1] >= 0) and (0 < top_layer[f][1] < 4) and not((top_layer[f][1] == 4-space[f]) and (top_layer[t][0] == -1))):
        return True
    else:
        return False

def transfer(glass, top_layer, space, f, t): # fからtに色水を移す
    n = top_layer[f][1]
    for _ in range(n):
        glass[t].appendleft(glass[f].popleft())
    top_layer[f] = topLayerCheck(glass, f)
    top_layer[t] = topLayerCheck(glass, t)
    space[f] = 4 - len(glass[f])
    space[t] = 4 - len(glass[t])
    return glass, top_layer, space

def completedCheck(top_layer): # 解けたかどうか確認
    global num_glass
    for i in range(num_glass):
        if(top_layer[i][1] != -1 and top_layer[i][1] != 4):
            return False
    return True

def printAns(ans): # 解答を標準出力
    print("ans:")
    for i in range(len(ans)):
        print(f"{ans[i][0]+1}->{ans[i][1]+1} : {ans[i][2]}")

def printGlass(glass): # グラスの状態を標準出力
    print("glass:")
    for i in range(len(glass)):
        print(f"{i} : {glass[i]}")

def printTopLayer(top_layer): # 一番上にある層の色と高さを出力
    print("top layer:")
    for i in range(len(glass)):
        print(f"{i} : {top_layer[i]}")

def makeGlassStr(glass, space): # グラスの状態を表す文字列を生成
    global num_glass
    s = ""
    a_glass = list(glass)
    for i in range(num_glass):
        n = len(a_glass[i])
        for _ in range(space[i]): s += '.'
        for j in range(n):
            t = a_glass[i][j]
            s += t
    return s        

def getColorCode(c): # 色を示す文字をカラーコードに変換
    if(c == 'r'): return 'ff69b4'
    elif(c == 'b'): return '1e90ff'
    elif(c == 'y'): return 'dddd00'
    elif(c == 'g'): return 'adff2f'
    elif(c == 'o'): return 'ffa500'
    elif(c == 'w'): return 'ffefd5'
    elif(c == 'd'): return 'b22222'
    elif(c == 'i'): return '4b0082'
    elif(c == 'B'): return '8b4513'
    elif(c == 'G'): return '008000'
    elif(c == 'v'): return 'dda0dd'
    elif(c == 'p'): return 'ffc0cb'
    else:           return '000000'

def cell(i,  j): # i行j列目のセルを返す
    global ws
    return ws.cell(row=i, column=j)

def clearAns(): # Excelの回答欄をリセット
    global wb, ws

    i = 2
    while(cell(i, 12).value != None):
        cell(i, 12).value = None
        cell(i, 13).value = None
        cell(i, 14).value = None
        cell(i, 15).fill = PatternFill()
        cell(i, 16).value = None
        i += 1

def printAns_toElsx(ans): # 解答をExcelに出力
    global wb, ws
    
    clearAns()
    for i in range(len(ans)):
        v = ans[i]
        cell(i+2, 12).value = v[0]+1
        cell(i+2, 13).value = "->"
        cell(i+2, 14).value = v[1]+1
        cell(i+2, 15).fill = PatternFill('solid', getColorCode(v[2]))
        cell(i+2, 16).value = v[3]

    wb.save("c:/Users/ochi yuma/OneDrive/ドキュメント/programming/Python/LoveWaterGlass.xlsm")
    os.startfile("c:/Users/ochi yuma/OneDrive/ドキュメント/programming/Python/LoveWaterGlass.xlsm")

def dfs(glass, top_layer, space, ans): # 解答を探す深さ優先探索
    global num_glass, num_color, appeared, wb

    printGlass(glass)

    if(makeGlassStr(glass, space) in appeared):
        return
    
    appeared.add(makeGlassStr(glass, space))

    if(completedCheck(top_layer)):
        printAns(ans)
        cell(1, 4).value = num_color
        cell(1, 8).value = num_glass
        printAns_toElsx(ans)
        print("close the Excel file")
        exit()

    queue = deque()
    for i in range(num_glass):
        for j in range(num_glass):
            if(i == j): continue
            if(transferableJudge(top_layer, space, i, j)): queue.append([i, j])

        while(len(queue)):
            ft = queue.popleft()
            ans.append([ft[0], ft[1], top_layer[ft[0]][0], top_layer[ft[0]][1]])

            # 状態をコピー
            glass_copy = copy.deepcopy(glass)
            top_layer_copy = copy.deepcopy(top_layer)
            space_copy = copy.deepcopy(space)

            # 実際に操作
            glass, top_layer, space = transfer(glass, top_layer, space, ft[0], ft[1])

            dfs(glass, top_layer, space, ans)

            # 状態を戻す
            glass = glass_copy
            top_layer = top_layer_copy
            space = space_copy

            ans.pop()  # 操作を取り消す（ansからも）

### 初期状態の入力 ###
print("red->r, blue->b, yellow->y, rightgreen->g, orange->o, white->w, deepred->d, indigo->i, brown->B, green->G, violet->v, pink->p")
print("glass first state:")
for i in range(num_color):
    glass.append(deque(list(map(str, input(f"{i+1} : ").split()))))
for i in range(num_glass - num_color):
    glass.append(deque())

### Excelで盤面を描画 ###
# グラスを空に
for i in range(3):
    for j in range(5):
        cell(6*(i+1)-3, 2*(j+1)+1).value = ""

# 色をリセットしてから入れる
lglass = list(glass)

if(num_color == 5):    r, c = 2, 4
elif(num_color == 7):  r, c = 2, 5
elif(num_color == 9):  r, c = 3, 4
elif(num_color == 12): r, c = 3, 5
else:                  r, c = 0, 0

l = 0
for i in range(r):
    for j in range(c):
        for k in range(4):
            ws[6*(i+1)-k][2*j+1].fill = PatternFill('solid', getColorCode(lglass[l][-k-1]))
        l += 1
        if(l == num_color): break
    if(l == num_color): break

### 最上層の確認 ###
for i in range(num_glass):
    top_layer.append(topLayerCheck(glass, i))

### 空き具合の記録 ###
for i in range(num_glass): space.append(4 - len(glass[i]))

### 解答 ###
dfs(glass, top_layer, space, ans)
